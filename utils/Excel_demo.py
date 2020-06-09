# -*- coding:utf-8 -*-
# @Time: 2020/6/9 10:56
# @Author: wenqin_zhu
# @File: Excel_demo.py
# @Software: PyCharm

# 导入加载workbook对象
import openpyxl

# TODO 方式一 通过active属性，打开当前活动表单
# filename = "cases.xlsx"
# # 加载并打开Excel
# my_workbook = openpyxl.load_workbook(filename=filename)
# # 打开当前活动的表单
# sheet = my_workbook.active
# # 定位到指定的单元格
# cell = sheet.cell(2, 3)
# # 通过value属性获取值
# get_c_value = cell.value
# print(get_c_value)
# # 查看读取出来的文本的数据类型
# print(type(get_c_value))

# get_c_value = tuple(cell.value)
# print(type(get_c_value))


# TODO 方式二 通过指定具体的表单名来打开对应的表单
# filename = "cases.xlsx"
# # 加载并打开Excel
# my_workbook = openpyxl.load_workbook(filename=filename)
# # 通过给出具体的表单名称
# sheet = my_workbook["Sheet1"]
# # 定位到指定的单元格
# cell = sheet.cell(2, 3)
# # 通过value属性获取值
# get_c_value = cell.value
# print(get_c_value)
# # 查看读取出来的文本的数据类型
# print(type(get_c_value))

# TODO 往单元格内写入数据。 打开 - 写入文本 - 保存
# filename = "cases.xlsx"
# # 1. 加载并打开Excel
# my_workbook = openpyxl.load_workbook(filename=filename)
# # 2. 通过给出具体的表单名称打开对应的表单
# sheet = my_workbook["Sheet1"]
# # 3. 定位到指定的单元格并写入目标内容文本
# cell = sheet.cell(2, 6, "Pass")
# # 4. 写完文本内容进行文件的保存
# my_workbook.save("cases.xlsx")


# TODO 实际工作中循环读取单元格内的所有数据
# Excel地址
# filename = "cases.xlsx"
# # 1. 加载并打开Excel
# my_workbook = openpyxl.load_workbook(filename=filename)
# # 2. 通过给出具体的表单名称打开对应的表单
# sheet = my_workbook["Sheet1"]
# 循环取出单元格内的文本 分别遍历行和列进行取值
# for row in range(sheet.min_row, sheet.max_row):
# Excel第一行为表头，从第二行开始遍历
# for row in range(2, sheet.max_row+1):
#     for col in range(sheet.min_column, sheet.max_column+1):
#         appoint_cell = sheet.cell(row=row, column=col)
#         value = appoint_cell.value
#         print(f"{value}-------{type(value)}")

# # 取出表头 行数为1   依次取出所有的列
# for col in range(sheet.min_column, sheet.max_column):
#     cell_value = sheet.cell(row=1, column=col)
#     print(cell_value.value)

# TODO 测试用例存在Excel中，执行测试用例之前先取出所有的测试用例数据
# 使用字典的数据格式接收测试数据    读取第一行表头作为字典的key值，大于等于1的行列数据为编写的测试用例
# 1、先定义字典
# case_dict = {}
# # 2、取出表头数据
# for col in range(sheet.min_column, sheet.max_column):
#     # 获取第一行每一列的值作为字典的key
#     cell_til = sheet.cell(row=1, column=col).value
#     case_dict[cell_til] = ""
# print(case_dict)
# # 3、取出除第一行以为的所有列的测试数据
# # 先定义列表存储数据
# for row in range(2, sheet.max_row+1):
#     for col in range(sheet.min_column, sheet.max_column+1):
#         cell_value = sheet.cell(row=row, column=col).value
#         print(cell_value)


# Excel地址
# filename = "cases.xlsx"
# 1. 加载并打开Excel
# my_workbook = openpyxl.load_workbook(filename=filename)
# 2. 通过给出具体的表单名称打开对应的表单
# my_sheet = my_workbook["Sheet1"]
# 通过表单的 iter_rows方法获取单元格数据
# 获取表头
# case_til = tuple(my_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=my_sheet.max_column, values_only=True))[0]
# print(case_til)
# # 获取表头之外的单元格数据
# case_data = tuple(my_sheet.iter_rows(min_row=2, min_col=1, values_only=True))
# for data in tuple(my_sheet.iter_rows(min_row=2, min_col=1, values_only=True)):
#     result = dict(zip(case_til, data))
#     print(result)

# for data in tuple(my_sheet.iter_rows(min_row=2, min_col=1, values_only=True)):
#     # dic = dict(map(lambda x, y: (x, y), case_til, data))
#     dic = lambda x, y: [x, y], case_til, data
#     # dic = dict(map(lambda x, y: (x, y), case_til, data))
#     # dic = dict(map(lambda x, y: (x, y), case_til, data))
#     print(dic)

# case_til = ('id', 'title', 'value', 'except', 'actual', 'assert_result')
# data = (1, 2, 3, 4, 5, 6)
#
#
# def a(x, y):
#     return x, y


# print(a(case_til, data))
# a = dict(zip(('id', 'title', 'value', 'except', 'actual', 'assert_result'), (1, 2, 3, 4, 5, 6)))
# print(a)
# lambda x, y: [x, y], case_til, data
