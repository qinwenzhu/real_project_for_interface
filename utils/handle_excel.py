# -*- coding:utf-8 -*-
# @Time: 2020/6/9 19:13
# @Author: wenqin_zhu
# @File: handle_excel.py
# @Software: PyCharm

import openpyxl


class HandleExcel:

    def __init__(self, file_path, sheet_name=None):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def open_file(self):
        """ 打开Excel文件
            无论是读取还是写入，都需要先打开目标文件 """

        # 打开excel表格
        load_workbook = openpyxl.load_workbook(filename=self.file_path)
        if self.sheet_name is None:
            # 坑比较大，会自动定位到Excel文件中的活动表单<如果Excel文件在关闭的时候，活动窗口为空的sheet表单，那么读取的excel表单数据为空>
            operation_sheet = load_workbook.active
        else:
            operation_sheet = load_workbook[sheet_name]
        return operation_sheet

    def read_all_data_from_excel(self):
        """ 读取指定表单内的全部的测试数据 """

        # 打开需要读取测试用例的文件目录
        operation_sheet = self.open_file()

        # 读取测试用例的表头
        case_til = tuple(operation_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=operation_sheet.max_column, values_only=True))[0]
        # print(case_til)
        # 定义列表来接收每一行的一条测试用例并追加到列表内
        result_list = []
        # 循环读取所有测试用例数据<除表头外>
        # print(tuple(self.operation_sheet.iter_rows(min_row=2, values_only=True)))
        for data in tuple(operation_sheet.iter_rows(min_row=2, values_only=True)):
            # case_data = dict(zip(case_til, data))
            # result_list.append(case_data)
            result_list.append(dict(zip(case_til, data)))
        return result_list

    def read_appoint_data_from_excel(self, row):
        """ 读取指定表单内的指定行的测试数据 """
        return self.read_all_data_from_excel()[row-1]

    def write_data_to_excel(self, **kwargs):

        # 打开需要读取测试用例的文件目录
        # operation_sheet = self.open_file()

        if kwargs is not None:
            for key, value in kwargs.items():
                print(key, value)
               # print(kwargs[key])
               # print(value)
            print(kwargs.items())

        # for key in kwargs.keys():
        #     pass

            # 数据写入
            # operation_sheet.cell(row=row, column=operation_sheet.max_column-2, value=args)


if __name__ == '__main__':

    from guard.tools.path import SharePath

    case_file = r"{}/device_case.xlsx".format(SharePath.DATA_FOLDER)
    sheet_name = "device"
    data = {"actual": 1, "result": 2}
    HandleExcel(case_file).write_data_to_excel(**data)
    # print(HandleExcel(case_file).read_all_data_from_excel())
    pass
